# Imports the Google Cloud client library
from google.cloud import translate

# Instantiaties a client
from openpyxl.utils import get_column_letter

translate_client = translate.Client()

# The text to translate
text = u'Hello, World!'
# The target language
target = 'ru'

# Translates some test into Russian
translation = translate_client.translate(
    text,
    target_language=target
)

print(u'Text: {}'.format(text))
print(u'Translation: {}'.format(translation['translatedText']))
##############################################################################

from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter

wb = Workbook()
dest_filename = 'empty_book.xlsx'

ws1 = wb.active
ws1.title = "range names"
for row in  range(1, 40):
    ws1.append(range(600))

ws2 = wb.create_sheet(title='Pi')
ws2['F5'] = 3.14

ws3 = wb.create_sheet(title='Data')
for row in range(10, 20):
    for col in range(27, 54):
        _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
print(ws3['AA10'].value)

wb.save(filename=dest_filename)

